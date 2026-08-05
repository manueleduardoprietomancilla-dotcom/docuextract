import os
import base64
import json
import sqlite3
import hashlib
import secrets
from datetime import datetime, timedelta
from fastapi import FastAPI, File, UploadFile, HTTPException, Header, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = FastAPI(title="DocuExtract AI")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")

ALLOWED_TYPES = {
    "application/pdf": "pdf",
    "image/jpeg": "jpeg",
    "image/jpg": "jpeg",
    "image/png": "png",
    "image/webp": "webp",
    "image/gif": "gif",
}

PLANS = {
    "free": {"name": "Free Trial", "limit": 3, "price_cop": 0, "link": ""},
    "starter": {"name": "Starter", "limit": 100, "price_cop": 37000, "link": "https://mpago.la/1aAxegd"},
    "pro": {"name": "Pro", "limit": 300, "price_cop": 78000, "link": "https://mpago.la/1UvoSPg"},
}

EXTRACTION_PROMPT = """Analyze this document carefully.

FIRST, determine which case applies:

CASE A — Multiple separate documents (different invoice numbers, different dates, different parties):
→ Return a JSON ARRAY, one object per document.

CASE B — One single document that spans multiple pages (same invoice number, same transaction, just long):
→ Return a single JSON OBJECT. Do NOT split it.

RULES to distinguish:
- If you see multiple different invoice/document numbers → CASE A
- If you see multiple different issuers or recipients → CASE A
- If the document is just one invoice/contract with many line items across pages → CASE B
- If unsure, treat as CASE B (single document)

For each document (whether one or many), use this exact structure:
{
  "document_type": "invoice|contract|receipt|quote|other",
  "language": "detected language",
  "summary": "brief one-line description",
  "key_fields": {
    "date": "document date if found",
    "due_date": "due date if found",
    "document_number": "invoice/document number if found",
    "total_amount": "total amount with currency",
    "subtotal": "subtotal if found",
    "tax": "tax amount if found",
    "currency": "currency code (USD, EUR, MXN, COP, etc)"
  },
  "parties": {
    "issuer": {
      "name": "company or person name",
      "address": "address if found",
      "tax_id": "tax ID / RFC / NIF / NIT if found",
      "email": "email if found",
      "phone": "phone if found"
    },
    "recipient": {
      "name": "company or person name",
      "address": "address if found",
      "tax_id": "tax ID / RFC / NIF / NIT if found"
    }
  },
  "line_items": [
    {
      "description": "item description",
      "quantity": "quantity",
      "unit_price": "unit price",
      "total": "line total"
    }
  ],
  "additional_info": {
    "payment_method": "payment method if found",
    "payment_terms": "payment terms if found",
    "notes": "any important notes"
  }
}

Return ONLY valid JSON. No markdown, no explanation. If a field is not found, use null."""


def get_db():
    db_path = os.environ.get("DB_PATH", "docuextract.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            plan TEXT DEFAULT 'free',
            docs_used INTEGER DEFAULT 0,
            docs_reset_date TEXT,
            token TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()
    conn.close()


init_db()


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def get_user_by_token(token: str):
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE token = ?", (token,)).fetchone()
    conn.close()
    return user


def reset_docs_if_needed(user):
    if not user["docs_reset_date"]:
        return
    reset_date = datetime.fromisoformat(user["docs_reset_date"])
    if datetime.now() > reset_date:
        conn = get_db()
        next_reset = (datetime.now() + timedelta(days=30)).isoformat()
        conn.execute("UPDATE users SET docs_used = 0, docs_reset_date = ? WHERE id = ?",
                     (next_reset, user["id"]))
        conn.commit()
        conn.close()


class RegisterRequest(BaseModel):
    email: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


class ActivateRequest(BaseModel):
    token: str
    plan: str


@app.get("/")
async def root():
    return FileResponse("static/index.html")


@app.post("/auth/register")
async def register(req: RegisterRequest):
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE email = ?", (req.email,)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=400, detail="Email already registered")
    token = secrets.token_hex(32)
    next_reset = (datetime.now() + timedelta(days=30)).isoformat()
    conn.execute(
        "INSERT INTO users (email, password_hash, token, plan, docs_reset_date) VALUES (?, ?, ?, 'free', ?)",
        (req.email, hash_password(req.password), token, next_reset)
    )
    conn.commit()
    conn.close()
    return {"success": True, "token": token, "email": req.email, "plan": "free", "docs_used": 0, "docs_limit": 3}


@app.post("/auth/login")
async def login(req: LoginRequest):
    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE email = ? AND password_hash = ?",
        (req.email, hash_password(req.password))
    ).fetchone()
    conn.close()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return {
        "success": True,
        "token": user["token"],
        "email": user["email"],
        "plan": user["plan"],
        "docs_used": user["docs_used"],
        "docs_limit": PLANS.get(user["plan"], {}).get("limit", 0)
    }


@app.get("/auth/me")
async def me(authorization: Optional[str] = Header(None)):
    token = authorization.replace("Bearer ", "") if authorization else None
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid token")
    reset_docs_if_needed(user)
    user = get_user_by_token(token)
    return {
        "email": user["email"],
        "plan": user["plan"],
        "docs_used": user["docs_used"],
        "docs_limit": PLANS.get(user["plan"], {}).get("limit", 0)
    }


@app.post("/auth/activate")
async def activate_plan(req: ActivateRequest):
    if req.plan not in PLANS:
        raise HTTPException(status_code=400, detail="Invalid plan")
    user = get_user_by_token(req.token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid token")
    conn = get_db()
    next_reset = (datetime.now() + timedelta(days=30)).isoformat()
    conn.execute(
        "UPDATE users SET plan = ?, docs_used = 0, docs_reset_date = ? WHERE token = ?",
        (req.plan, next_reset, req.token)
    )
    conn.commit()
    conn.close()
    return {"success": True, "plan": req.plan}


@app.get("/plans")
async def get_plans():
    return {"plans": PLANS}


@app.post("/extract")
async def extract_document(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
    x_api_key: Optional[str] = Header(None)
):
    token = authorization.replace("Bearer ", "") if authorization else None
    user = get_user_by_token(token) if token else None

    if user:
        reset_docs_if_needed(user)
        user = get_user_by_token(token)
        plan_limit = PLANS.get(user["plan"], {}).get("limit", 0)
        if plan_limit == 0:
            raise HTTPException(status_code=403, detail="Please subscribe to a plan to process documents.")
        if user["docs_used"] >= plan_limit:
            raise HTTPException(status_code=403, detail=f"Monthly limit reached ({plan_limit} documents). Please upgrade your plan.")
        api_key = os.environ.get("ANTHROPIC_API_KEY")
    else:
        api_key = x_api_key or os.environ.get("ANTHROPIC_API_KEY")

    if not api_key:
        raise HTTPException(status_code=400, detail="API key required.")

    content_type = file.content_type
    if content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="File type not supported. Use PDF, JPG, PNG, WEBP or GIF.")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB.")

    file_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
    media_type = content_type if content_type != "image/jpg" else "image/jpeg"

    client = anthropic.Anthropic(api_key=api_key)

    if content_type == "application/pdf":
        message = client.messages.create(
            model="claude-sonnet-5",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": file_b64}},
                    {"type": "text", "text": EXTRACTION_PROMPT}
                ]
            }]
        )
    else:
        message = client.messages.create(
            model="claude-sonnet-5",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": file_b64}},
                    {"type": "text", "text": EXTRACTION_PROMPT}
                ]
            }]
        )

    raw_text = message.content[0].text.strip()

    # Clean markdown code blocks if present
    if "```" in raw_text:
        import re
        match = re.search(r'```(?:json)?\s*([\s\S]*?)```', raw_text)
        if match:
            raw_text = match.group(1).strip()
        else:
            lines = raw_text.split("\n")
            raw_text = "\n".join(l for l in lines if not l.strip().startswith("```"))

    # Find JSON content (object or array) in case there's extra text
    raw_text = raw_text.strip()
    if not (raw_text.startswith("{") or raw_text.startswith("[")):
        import re
        match = re.search(r'(\{[\s\S]*\}|\[[\s\S]*\])', raw_text)
        if match:
            raw_text = match.group(1)

    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Could not parse AI response. Please try again. Error: {str(e)}")

    if isinstance(parsed, list):
        documents = parsed
    else:
        documents = [parsed]

    if user:
        conn = get_db()
        conn.execute("UPDATE users SET docs_used = docs_used + 1 WHERE token = ?", (token,))
        conn.commit()
        conn.close()

    return {"success": True, "data": documents[0] if len(documents) == 1 else documents, "multiple": len(documents) > 1, "count": len(documents), "filename": file.filename}


@app.post("/export-excel")
async def export_excel(data: dict):
    raw = data.get("data", {})
    filename = data.get("filename", "document")
    lang = data.get("lang", "en")

    LABELS = {
        "en": {
            "doc_info": "DOCUMENT INFO", "doc_type": "Document Type", "language": "Language", "summary": "Summary",
            "key_fields": "KEY FIELDS", "issuer": "ISSUER / FROM", "recipient": "RECIPIENT / TO",
            "line_items": "LINE ITEMS", "additional": "ADDITIONAL INFO",
            "desc": "Description", "qty": "Qty", "unit": "Unit Price", "total": "Total",
        },
        "es": {
            "doc_info": "INFORMACIÓN DEL DOCUMENTO", "doc_type": "Tipo de Documento", "language": "Idioma", "summary": "Resumen",
            "key_fields": "CAMPOS CLAVE", "issuer": "EMISOR / DE", "recipient": "RECEPTOR / PARA",
            "line_items": "ARTÍCULOS", "additional": "INFORMACIÓN ADICIONAL",
            "desc": "Descripción", "qty": "Cant.", "unit": "Precio Unit.", "total": "Total",
        },
        "pt": {
            "doc_info": "INFORMAÇÃO DO DOCUMENTO", "doc_type": "Tipo de Documento", "language": "Idioma", "summary": "Resumo",
            "key_fields": "CAMPOS CHAVE", "issuer": "EMISSOR / DE", "recipient": "DESTINATÁRIO / PARA",
            "line_items": "ITENS", "additional": "INFORMAÇÃO ADICIONAL",
            "desc": "Descrição", "qty": "Qtd.", "unit": "Preço Unit.", "total": "Total",
        },
        "fr": {
            "doc_info": "INFORMATIONS DU DOCUMENT", "doc_type": "Type de Document", "language": "Langue", "summary": "Résumé",
            "key_fields": "CHAMPS CLÉS", "issuer": "ÉMETTEUR / DE", "recipient": "DESTINATAIRE / À",
            "line_items": "ARTICLES", "additional": "INFORMATIONS SUPPLÉMENTAIRES",
            "desc": "Description", "qty": "Qté", "unit": "Prix Unit.", "total": "Total",
        },
    }
    L = LABELS.get(lang, LABELS["en"])
    # Support both single doc and multiple docs
    docs = raw if isinstance(raw, list) else [raw]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    section_font = Font(color="4fc3f7", bold=True, size=10)

    def write_sheet(ws, extracted, doc_label):
        def write_header(row, text):
            cell = ws.cell(row=row, column=1, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(f"A{row}:D{row}")

        def write_section(row, text):
            cell = ws.cell(row=row, column=1, value=text)
            cell.fill = section_fill
            cell.font = section_font
            ws.merge_cells(f"A{row}:D{row}")

        def write_row(row, label, value):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value) if value else "")

        r = 1
        write_header(r, f"DocuExtract AI - {doc_label}")
        r += 2

        write_section(r, L["doc_info"]); r += 1
        write_row(r, L["doc_type"], extracted.get("document_type", "")); r += 1
        write_row(r, L["language"], extracted.get("language", "")); r += 1
        write_row(r, L["summary"], extracted.get("summary", "")); r += 2

        kf = extracted.get("key_fields", {})
        if kf:
            write_section(r, L["key_fields"]); r += 1
            for k, v in kf.items():
                if v:
                    write_row(r, k.replace("_", " ").title(), v); r += 1
            r += 1

        parties = extracted.get("parties", {})
        issuer = parties.get("issuer", {})
        recipient = parties.get("recipient", {})

        if any(v for v in issuer.values() if v):
            write_section(r, L["issuer"]); r += 1
            for k, v in issuer.items():
                if v:
                    write_row(r, k.replace("_", " ").title(), v); r += 1
            r += 1

        if any(v for v in recipient.values() if v):
            write_section(r, L["recipient"]); r += 1
            for k, v in recipient.items():
                if v:
                    write_row(r, k.replace("_", " ").title(), v); r += 1
            r += 1

        items = extracted.get("line_items", [])
        if items:
            write_section(r, L["line_items"]); r += 1
            for col, h in enumerate([L["desc"], L["qty"], L["unit"], L["total"]], 1):
                ws.cell(row=r, column=col, value=h).font = Font(bold=True)
            r += 1
            for item in items:
                ws.cell(row=r, column=1, value=item.get("description", ""))
                ws.cell(row=r, column=2, value=item.get("quantity", ""))
                ws.cell(row=r, column=3, value=item.get("unit_price", ""))
                ws.cell(row=r, column=4, value=item.get("total", ""))
                r += 1
            r += 1

        additional = extracted.get("additional_info", {})
        if any(v for v in additional.values() if v):
            write_section(r, L["additional"]); r += 1
            for k, v in additional.items():
                if v:
                    write_row(r, k.replace("_", " ").title(), v); r += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

    for i, doc in enumerate(docs):
        sheet_title = f"Doc {i+1}" if len(docs) > 1 else "Extracted Data"
        ws = wb.create_sheet(title=sheet_title)
        label = f"{filename} ({i+1}/{len(docs)})" if len(docs) > 1 else filename
        write_sheet(ws, doc, label)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    safe_name = filename.rsplit(".", 1)[0] if "." in filename else filename
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_extracted.xlsx"},
    )
